import pytest
import os
import sys
from datetime import datetime
from playwright.sync_api import sync_playwright
import allure
import openpyxl
from openpyxl import Workbook
from pages.login_page import LoginPage
from utils.config import BASE_URL, USERNAME, PASSWORD

# Global list to store test results for Excel report
test_results_data = []

def pytest_collection_modifyitems(config, items):
    """Enforce test execution priority: Login -> Job Actions -> Job Category"""
    # Priority order for test files
    order = ["test_login.py", "test_user_management.py", "test_job_actions.py", "test_employment_status.py", "test_job_categories.py", "test_work_shifts.py", "test_general_info.py", "test_locations.py" , "test_organization_structure.py"]
    
    def get_order_priority(item):
        filename = os.path.basename(item.fspath)
        try:
            return order.index(filename)
        except ValueError:
            return len(order)

    # Sort items based on the priority list
    items.sort(key=get_order_priority)

@pytest.fixture(scope="session")
def playwright_instance():
    with sync_playwright() as p:
        yield p

@pytest.fixture(scope="session")
def browser(playwright_instance):
    browser = playwright_instance.chromium.launch(headless=False)
    yield browser
    browser.close()

@pytest.fixture(scope="session")
def session_context(browser):
    context = browser.new_context()
    yield context
    context.close()

@pytest.fixture(scope="session")
def session_page(session_context):
    page = session_context.new_page()
    yield page
    page.close()

@pytest.fixture(scope="session")
def authenticated_page(session_page):
    """Use the shared session page and sign in once."""
    if "dashboard/index" not in session_page.url:
        login_page = LoginPage(session_page)
        # Only navigate if not already on login page
        if "/login" not in session_page.url:
            login_page.navigate(BASE_URL)
        
        login_page.login(USERNAME, PASSWORD)
        
        # Wait for dashboard to load
        session_page.wait_for_url("**/dashboard/index", timeout=60000)
    
    yield session_page

@pytest.fixture(scope="function")
def context(browser):
    context = browser.new_context()
    yield context
    context.close()

@pytest.fixture(scope="function")
def page(context):
    page = context.new_page()
    yield page
    page.close()

@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report = outcome.get_result()

    if report.when == "call":
        # Capture data for Excel report
        test_id = getattr(item, "cell_id", "N/A") # Placeholder for TC ID if used
        test_name = item.name
        expected_result = item.obj.__doc__.strip() if item.obj.__doc__ else "No description provided"
        actual_result = "Test Passed Successfully" if report.passed else str(report.longreprtext)
        status = report.outcome.upper()

        test_results_data.append({
            "id": test_id,
            "name": test_name,
            "expected": expected_result,
            "actual": actual_result,
            "status": status
        })

        if report.failed:
            # Try to get page from authenticated_page or page fixture
            page = item.funcargs.get("authenticated_page") or item.funcargs.get("page")
            if page:
                try:
                    screenshot = page.screenshot()
                    allure.attach(
                        screenshot,
                        name=f"{item.name}_failure",
                        attachment_type=allure.attachment_type.PNG
                    )
                except Exception as e:
                    print(f"Failed to take screenshot: {e}")

def pytest_sessionfinish(session, exitstatus):
    """Generate Excel report at the end of the test session."""
    if not test_results_data:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    # Define Headers
    headers = ["Testcase ID", "Testcases", "Expected Result", "Actual Result", "Status"]
    ws.append(headers)

    # Style headers
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Add data
    for result in test_results_data:
        ws.append([
            result["id"],
            result["name"],
            result["expected"],
            result["actual"][:1000], # Truncate very long error messages for Excel
            result["status"]
        ])

    # Basic column width adjustment
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)

    # Create reports directory if it doesn't exist
    reports_dir = os.path.join(session.config.rootdir, "reports")
    os.makedirs(reports_dir, exist_ok=True)

    report_path = os.path.join(reports_dir, "test_results.xlsx")
    wb.save(report_path)
    print(f"\nExcel report generated at: {report_path}")

    # Also generate CSV report for easier viewing in IDE
    import csv
    csv_path = os.path.join(reports_dir, "test_results.csv")
    with open(csv_path, mode='w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for result in test_results_data:
            writer.writerow([
                result["id"],
                result["name"],
                result["expected"],
                result["actual"].replace('\n', ' ')[:1000],
                result["status"]
            ])
    print(f"CSV report generated at: {csv_path}")
